# Imports FastAPI pour les routes et les erreurs
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from typing import List, Optional
from enum import Enum
from app.models import User, Employee, Bonus, Validation, PrimeMax, AuditLog, Notification, ValidationStatus
from app.auth import get_current_user
from app.schemas import *
from fastapi import HTTPException
import io
import csv
from datetime import datetime
from tortoise.expressions import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from decimal import Decimal


def _sanitize_for_json(v):
    if isinstance(v, Decimal):
        return str(v)
    if isinstance(v, Enum):
        return v.value if hasattr(v, 'value') else str(v)
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, dict):
        return {k: _sanitize_for_json(v) for k, v in v.items()}
    if isinstance(v, list):
        return [_sanitize_for_json(x) for x in v]
    return v

router = APIRouter(dependencies=[Depends(get_current_user)])

# Route POST pour créer une prime
@router.post("/bonuses/", response_model=BonusResponse)
async def create_bonus(bonus: BonusCreate, user: User = Depends(get_current_user)):
    employee = await Employee.get(id=bonus.employee_id)

    if bonus.bonus_type != BonusType.ASTREINTE:
        primemax = await PrimeMax.filter(
            dept_str=employee.dept_str,
            bonus_type=bonus.bonus_type
        ).first()
        # Le plafond ne s'applique qu'à l'évaluation (quanti + quali), pas aux "Autres primes"
        details = bonus.details or {}
        others_list = details.get('others', []) if isinstance(details, dict) else []
        others_total = sum(float(o.get('montant', 0) or 0) for o in others_list)
        eval_amount = float(bonus.total_amount) - others_total
        if primemax and eval_amount > primemax.amount:
            raise HTTPException(
                status_code=400,
                detail=f"Le montant de l'évaluation ({eval_amount} Ar) dépasse le plafond "
                       f"autorisé ({primemax.amount} Ar) pour "
                       f"'{bonus.bonus_type.value}' dans le département '{employee.dept_str}'."
            )

    existing = await Bonus.filter(
        employee_id=bonus.employee_id,
        bonus_type=bonus.bonus_type,
        start_date__lte=bonus.end_date,
        end_date__gte=bonus.start_date,
    ).exists()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"Une prime de type '{bonus.bonus_type.value}' existe déjà sur cette période pour cet employé."
        )
    initial_status = ValidationStatus.EN_ATTENTE_DIRECTEUR if user.is_directeur else ValidationStatus.INITIALISE
    obj = await Bonus.create(**bonus.dict(), created_by_id=user.id, status=initial_status)
    return await Bonus.get(id=obj.id).prefetch_related('employee')

# Route POST pour validation par lot
@router.post("/bonuses/batch/validate", response_model=BatchValidateResponse)
async def batch_validate_bonuses(
    request: BatchValidateRequest,
    user: User = Depends(get_current_user)
):
    results = []
    for bonus_id in request.bonus_ids:
        try:
            bonus = await Bonus.get_or_none(id=bonus_id)
            if not bonus:
                results.append(BatchValidateResult(bonus_id=bonus_id, success=False, error="Prime introuvable"))
                continue

            if bonus.status == ValidationStatus.VALIDE:
                results.append(BatchValidateResult(bonus_id=bonus_id, success=False, error="Déjà validée"))
                continue

            expected_status = {
                "N1": ValidationStatus.INITIALISE,
                "DIRECTEUR": ValidationStatus.EN_ATTENTE_DIRECTEUR,
                "DG": ValidationStatus.EN_ATTENTE_DG,
            }.get(request.step)

            if not expected_status:
                results.append(BatchValidateResult(bonus_id=bonus_id, success=False, error="Étape invalide"))
                continue

            if bonus.status != expected_status:
                results.append(BatchValidateResult(
                    bonus_id=bonus_id, success=False,
                    error=f"Statut actuel '{bonus.status}', attendait '{expected_status}'"
                ))
                continue

            # Vérifier que l'utilisateur a le rôle correspondant à l'étape
            allowed_roles = {
                "N1": user.is_validator_n1,
                "DIRECTEUR": user.is_directeur,
                "DG": user.is_dg,
            }
            if not allowed_roles.get(request.step, False):
                results.append(BatchValidateResult(bonus_id=bonus_id, success=False, error="Rôle non autorisé pour cette étape"))
                continue

            # Vérifier le département (sauf DG)
            if not user.is_dg:
                bonus_emp = await Employee.get_or_none(id=bonus.employee_id)
                if bonus_emp and bonus_emp.dept_str != user.department:
                    results.append(BatchValidateResult(bonus_id=bonus_id, success=False, error="Prime d'un autre département"))
                    continue

            await Validation.create(
                bonus_id=bonus.id,
                validator_id=user.id,
                step=request.step,
                action=request.action,
                note=request.note,
                motif_rejet=request.motif_rejet,
            )

            if request.action == "VALIDER":
                bonus.status = {
                    "N1": ValidationStatus.EN_ATTENTE_DIRECTEUR,
                    "DIRECTEUR": ValidationStatus.EN_ATTENTE_DG,
                    "DG": ValidationStatus.VALIDE
                }[request.step]

                if bonus.status == ValidationStatus.VALIDE:
                    await Validation.create(
                        bonus_id=bonus.id,
                        validator_id=user.id,
                        step="CLOSED",
                        action="AUTOMATIC",
                        note="Prime validée par DG - Clôture automatique"
                    )
            elif request.action == "REJETER":
                bonus.status = ValidationStatus.INITIALISE
                bonus.was_rejected = True

            await bonus.save()
            results.append(BatchValidateResult(bonus_id=bonus_id, success=True))

        except Exception as e:
            results.append(BatchValidateResult(bonus_id=bonus_id, success=False, error=str(e)))

    total_success = sum(1 for r in results if r.success)
    total_errors = len(results) - total_success
    return BatchValidateResponse(results=results, total_success=total_success, total_errors=total_errors)

# Route PUT pour modifier une prime
@router.put("/bonuses/{bonus_id}", response_model=BonusResponse)
async def update_bonus(bonus_id: int, data: BonusCreate, user: User = Depends(get_current_user)):
    bonus = await Bonus.get_or_none(id=bonus_id).prefetch_related('employee')
    if not bonus: raise HTTPException(404, "Bonus not found")

    # Vérifier le département (sauf DG)
    if not user.is_dg and bonus.employee.dept_str != user.department:
        raise HTTPException(status_code=403, detail="Vous ne pouvez pas modifier une prime d'un autre département")

    # Vérifier le rôle et le statut
    can_edit = False
    if user.is_dg and bonus.status == ValidationStatus.EN_ATTENTE_DG:
        can_edit = True
    elif user.is_directeur and bonus.status == ValidationStatus.EN_ATTENTE_DIRECTEUR:
        can_edit = True
    elif user.is_validator_n1 and bonus.status == ValidationStatus.INITIALISE:
        can_edit = True
    
    if not can_edit:
        raise HTTPException(status_code=403, detail="Vous n'êtes pas autorisé à modifier cette prime")

    update_data = data.dict(exclude_unset=True)

    # Si DG ou Directeur modifie, la prime revient à Initialisé pour re-validation
    if (user.is_dg or user.is_directeur) and bonus.status != ValidationStatus.INITIALISE:
        update_data['status'] = ValidationStatus.INITIALISE

    update_data['was_rejected'] = False

    SCALAR_FIELDS = [
        "total_amount", "performance_score", "absences", "retard",
        "prime_mensuel_amount", "nb_jours_astreinte", "taux_jour",
        "prime_astreinte_amount", "ca_realise", "ca_objectif",
        "taux_commission", "commission_amount",
    ]

    before = {f: getattr(bonus, f) for f in SCALAR_FIELDS}
    before["details"] = bonus.details
    before["status"] = str(bonus.status.value) if hasattr(bonus.status, 'value') else str(bonus.status)

    await bonus.update_from_dict(update_data)
    await bonus.save()
    updated = await Bonus.get(id=bonus.id).prefetch_related('employee')

    after = {f: getattr(updated, f) for f in SCALAR_FIELDS}
    after["details"] = updated.details
    after["status"] = str(updated.status.value) if hasattr(updated.status, 'value') else str(updated.status)

    changes = []
    for f in SCALAR_FIELDS:
        if str(before[f]) != str(after[f]):
            changes.append(f"{f}: {before[f]} → {after[f]}")
    if before["status"] != after["status"]:
        changes.append(f"statut: {before['status']} → {after['status']}")
    if str(before["details"]) != str(after["details"]):
        changes.append("détails modifiés (critères, notes, coefficients)")

    changes_data = {"before": {f: _sanitize_for_json(before[f]) for f in SCALAR_FIELDS + ["details", "status"] if str(before[f]) != str(after[f])},
                    "after": {f: _sanitize_for_json(after[f]) for f in SCALAR_FIELDS + ["details", "status"] if str(before[f]) != str(after[f])}}
    try:
        await AuditLog.create(
            bonus_id=bonus.id,
            user_id=user.id,
            action="MODIFICATION",
            description="; ".join(changes) if changes else "Modification enregistrée",
            changes=changes_data,
        )
    except Exception:
        pass  # Audit log failure must not block the modification

    if changes:
        employee = await Employee.get(id=bonus.employee_id).prefetch_related('manager')
        recipients = []

        if user.is_dg:
            if employee.manager:
                recipients.append(employee.manager)
        elif user.is_directeur:
            if employee.manager:
                recipients.append(employee.manager)

        FIELD_LABELS_SHORT = {
            "total_amount": "montant", "performance_score": "score", "status": "statut",
            "prime_mensuel_amount": "prime", "prime_astreinte_amount": "astreinte",
            "commission_amount": "commission", "nb_jours_astreinte": "jours astreinte",
            "taux_jour": "taux/jour", "taux_commission": "taux commission",
            "ca_realise": "CA réalisé", "ca_objectif": "CA objectif",
            "details": "détails", "absences": "absences", "retard": "retards",
        }
        changed_fields = set()
        for c in changes:
            field = c.split(":")[0]
            changed_fields.add(FIELD_LABELS_SHORT.get(field, field))
        summary = ", ".join(sorted(changed_fields))

        for r in recipients:
            try:
                await Notification.create(
                    user=r,
                    bonus=bonus,
                    sender=user,
                    type="MODIF_DG" if user.is_dg else "MODIF_DIR",
                    message=f"{employee.name} — {summary}",
                )
            except Exception:
                pass

    return updated

# Route GET pour lister les primes (filtres optionnels)
@router.get("/bonuses/", response_model=List[BonusResponse])
async def list_bonuses(
    status: Optional[str] = None,
    employee_id: Optional[int] = None,
    bonus_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    was_rejected: Optional[bool] = None,
    show_paid: Optional[bool] = False,
    all_statuses: Optional[bool] = False,
    view_all_validated: Optional[bool] = False,
    validated_by_me: Optional[bool] = False,
    user: User = Depends(get_current_user),
):
    query = Bonus.all().prefetch_related('employee')
    if not (user.is_dg or user.is_drh) and user.department:
        query = query.filter(employee__dept_str=user.department)

    # Filtrer les statuts selon le rôle de l'utilisateur (sauf si all_statuses pour Kanban)
    if user.is_dg:
        allowed_statuses = [ValidationStatus.EN_ATTENTE_DG]
    elif user.is_drh:
        allowed_statuses = [ValidationStatus.VALIDE]
    elif user.is_directeur:
        allowed_statuses = [ValidationStatus.EN_ATTENTE_DIRECTEUR]
    elif user.is_validator_n1:
        allowed_statuses = [ValidationStatus.INITIALISE]
    else:
        allowed_statuses = []

    if all_statuses:
        pass  # Kanban : toutes les primes du département, tous statuts
    elif validated_by_me:
        validated_ids = await Validation.filter(
            validator_id=user.id, action="VALIDER"
        ).values_list("bonus_id", flat=True)
        query = query.filter(id__in=validated_ids, status=ValidationStatus.VALIDE)
    elif view_all_validated:
        query = query.filter(status=ValidationStatus.VALIDE)  # Page validées DG/DRH
    elif status:
        status_enum = ValidationStatus(status)
        if status_enum not in allowed_statuses:
            raise HTTPException(status_code=403, detail="Vous n'avez pas accès aux primes avec ce statut")
        query = query.filter(status=status_enum)
    else:
        query = query.filter(status__in=allowed_statuses)

    if show_paid:
        query = query.filter(paid_at__isnull=False)
    else:
        query = query.filter(paid_at__isnull=True)
    if employee_id: query = query.filter(employee_id=employee_id)
    if bonus_type: query = query.filter(bonus_type=bonus_type)
    if start_date: query = query.filter(start_date__gte=start_date)
    if end_date: query = query.filter(end_date__lte=end_date)
    if was_rejected is not None: query = query.filter(was_rejected=was_rejected)
    return await query


@router.get("/bonuses/export")
async def export_bonuses(
    status: Optional[str] = None,
    employee_id: Optional[int] = None,
    bonus_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    department: Optional[str] = None,
    search: Optional[str] = None,
    columns: Optional[str] = None,
    was_rejected: Optional[bool] = None,
    show_paid: Optional[bool] = False,
    user: User = Depends(get_current_user),
):
    query = Bonus.all().prefetch_related('employee', 'created_by')

    # Filtre département selon le rôle
    if not (user.is_dg or user.is_drh) and user.department:
        query = query.filter(employee__dept_str=user.department)

    # Filtre statut selon le rôle
    if user.is_dg:
        allowed_statuses = [ValidationStatus.EN_ATTENTE_DG]
    elif user.is_drh:
        allowed_statuses = [ValidationStatus.VALIDE]
    elif user.is_directeur:
        allowed_statuses = [ValidationStatus.EN_ATTENTE_DIRECTEUR]
    elif user.is_validator_n1:
        allowed_statuses = [ValidationStatus.INITIALISE]
    else:
        allowed_statuses = []

    if status:
        status_enum = ValidationStatus(status)
        if status_enum not in allowed_statuses:
            raise HTTPException(status_code=403, detail="Vous n'avez pas accès aux primes avec ce statut")
        query = query.filter(status=status_enum)
    else:
        query = query.filter(status__in=allowed_statuses)

    if employee_id: query = query.filter(employee_id=employee_id)
    if bonus_type: query = query.filter(bonus_type=bonus_type)
    if start_date and end_date:
        query = query.filter(start_date__lte=end_date, end_date__gte=start_date)
    elif start_date:
        query = query.filter(start_date__gte=start_date)
    elif end_date:
        query = query.filter(end_date__lte=end_date)
    if department:
        if not (user.is_dg or user.is_drh) and department != user.department:
            raise HTTPException(status_code=403, detail="Vous ne pouvez exporter que les primes de votre département")
        query = query.filter(employee__dept_str=department)
    if was_rejected is not None: query = query.filter(was_rejected=was_rejected)
    if search:
        query = query.filter(
            Q(employee__name__icontains=search) | Q(employee__matricule__icontains=search)
        )
    if show_paid:
        query = query.filter(paid_at__isnull=False)

    bonuses = await query.order_by('-start_date')

    all_columns = [
        "Matricule", "Nom", "Departement", "TypePrime",
        "DateDebut", "DateFin", "Montant", "Statut",
        "DejaRejete", "MarqueePayeeLe", "CreePar", "DateCreation"
    ]
    if columns:
        selected = [c.strip() for c in columns.split(',') if c.strip() in all_columns]
    else:
        selected = all_columns[:]

    extractors = {
        "Matricule": lambda b: b.employee.matricule,
        "Nom": lambda b: b.employee.name,
        "Departement": lambda b: b.employee.department if b.employee.department else '',
        "TypePrime": lambda b: b.bonus_type.value,
        "DateDebut": lambda b: b.start_date.isoformat(),
        "DateFin": lambda b: b.end_date.isoformat(),
        "Montant": lambda b: str(int(b.total_amount)),
        "Statut": lambda b: b.status.value,
        "DejaRejete": lambda b: "Oui" if b.was_rejected else "Non",
        "MarqueePayeeLe": lambda b: b.paid_at.strftime('%d/%m/%Y %H:%M') if b.paid_at else '',
        "CreePar": lambda b: b.created_by.name if b.created_by else '',
        "DateCreation": lambda b: b.created_at.isoformat() if b.created_at else '',
    }

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(selected)
    for b in bonuses:
        writer.writerow([extractors[col](b) for col in selected])

    output.seek(0)
    return StreamingResponse(
        iter(['\ufeff' + output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=export_primes_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


@router.get("/bonuses/export/xlsx")
async def export_bonuses_xlsx(
    status: Optional[str] = None,
    employee_id: Optional[int] = None,
    bonus_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    department: Optional[str] = None,
    search: Optional[str] = None,
    columns: Optional[str] = None,
    was_rejected: Optional[bool] = None,
    user: User = Depends(get_current_user),
):
    query = Bonus.all().prefetch_related('employee', 'created_by')

    # Filtre département selon le rôle
    if not (user.is_dg or user.is_drh) and user.department:
        query = query.filter(employee__dept_str=user.department)

    # Filtre statut selon le rôle
    if user.is_dg:
        allowed_statuses = [ValidationStatus.EN_ATTENTE_DG]
    elif user.is_drh:
        allowed_statuses = [ValidationStatus.VALIDE]
    elif user.is_directeur:
        allowed_statuses = [ValidationStatus.EN_ATTENTE_DIRECTEUR]
    elif user.is_validator_n1:
        allowed_statuses = [ValidationStatus.INITIALISE]
    else:
        allowed_statuses = []

    if status:
        status_enum = ValidationStatus(status)
        if status_enum not in allowed_statuses:
            raise HTTPException(status_code=403, detail="Vous n'avez pas accès aux primes avec ce statut")
        query = query.filter(status=status_enum)
    else:
        query = query.filter(status__in=allowed_statuses)

    if employee_id: query = query.filter(employee_id=employee_id)
    if bonus_type: query = query.filter(bonus_type=bonus_type)
    if start_date and end_date:
        query = query.filter(start_date__lte=end_date, end_date__gte=start_date)
    elif start_date:
        query = query.filter(start_date__gte=start_date)
    elif end_date:
        query = query.filter(end_date__lte=end_date)
    if department:
        if not (user.is_dg or user.is_drh) and department != user.department:
            raise HTTPException(status_code=403, detail="Vous ne pouvez exporter que les primes de votre département")
        query = query.filter(employee__dept_str=department)
    if was_rejected is not None: query = query.filter(was_rejected=was_rejected)
    if search:
        query = query.filter(
            Q(employee__name__icontains=search) | Q(employee__matricule__icontains=search)
        )

    bonuses = await query.order_by('-start_date')

    all_columns = [
        "Matricule", "Nom", "Departement", "TypePrime",
        "DateDebut", "DateFin", "Montant", "Statut",
        "DejaRejete", "MarqueePayeeLe", "CreePar", "DateCreation"
    ]
    if columns:
        selected = [c.strip() for c in columns.split(',') if c.strip() in all_columns]
    else:
        selected = all_columns[:]

    extractors = {
        "Matricule": lambda b: b.employee.matricule,
        "Nom": lambda b: b.employee.name,
        "Departement": lambda b: b.employee.department,
        "TypePrime": lambda b: b.bonus_type.value,
        "DateDebut": lambda b: b.start_date.isoformat(),
        "DateFin": lambda b: b.end_date.isoformat(),
        "Montant": lambda b: b.total_amount,
        "Statut": lambda b: b.status.value,
        "DejaRejete": lambda b: "Oui" if b.was_rejected else "Non",
        "MarqueePayeeLe": lambda b: b.paid_at.strftime('%d/%m/%Y %H:%M') if b.paid_at else '',
        "CreePar": lambda b: b.created_by.name if b.created_by else '',
        "DateCreation": lambda b: b.created_at.isoformat() if b.created_at else '',
    }

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    number_font = Font(size=11)
    alt_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    dept_groups = {}
    for b in bonuses:
        dept = b.employee.department if b.employee.department else 'N/A'
        if dept not in dept_groups:
            dept_groups[dept] = []
        dept_groups[dept].append(b)

    def write_sheet(ws, title, data):
        ws.title = title
        for col_idx, col_name in enumerate(selected, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for row_idx, b in enumerate(data, 2):
            for col_idx, col_name in enumerate(selected, 1):
                val = extractors[col_name](b)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = number_font
                if col_name == "Montant" and isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
        for col_idx in range(1, len(selected) + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ''))
            for row_idx in range(2, min(len(data) + 2, 50)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)

    ws_all = wb.active
    write_sheet(ws_all, "Toutes", bonuses)

    for dept in sorted(dept_groups.keys()):
        ws = wb.create_sheet()
        write_sheet(ws, dept[:31], dept_groups[dept])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=export_primes_{datetime.now().strftime('%Y%m%d')}.xlsx"
        }
    )


@router.get("/bonuses/export/sage")
async def export_sage():
    bonuses = await Bonus.filter(status=ValidationStatus.VALIDE).prefetch_related('employee')

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow([
        "Matricule", "Nom", "Departement", "TypePrime",
        "DateDebut", "DateFin", "Montant", "Statut"
    ])
    for b in bonuses:
        writer.writerow([
            b.employee.matricule,
            b.employee.name,
            b.employee.department,
            b.bonus_type.value,
            b.start_date.isoformat(),
            b.end_date.isoformat(),
            str(int(b.total_amount)),
            b.status.value
        ])

    output.seek(0)
    return StreamingResponse(
        iter(['\ufeff' + output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=export_sage_paie.csv"}
    )


# Route GET pour l'export d'une prime spécifique
@router.get("/bonuses/{bonus_id}/export")
async def export_bonus_detail(bonus_id: int, columns: Optional[str] = None):
    bonus = await Bonus.get_or_none(id=bonus_id).prefetch_related('employee', 'created_by')
    if not bonus:
        raise HTTPException(404, "Bonus not found")

    common = [
        "Matricule", "Nom", "Departement", "TypePrime",
        "DateDebut", "DateFin", "MontantTotal", "Statut",
        "DejaRejete", "CreePar", "DateCreation"
    ]
    type_cols = {
        'mensuel': ["Score", "Quantitatif", "Qualitatif"],
        'astreinte': ["NbDisponibilite", "TotalDisponibilite", "TotalInterventions", "Exceptionnelle", "Ponctuelle"],
        'commission': ["NbVentes"],
    }
    all_possible = common + type_cols.get(bonus.bonus_type.value, [])

    if columns:
        selected = [c.strip() for c in columns.split(',') if c.strip() in all_possible]
    else:
        selected = all_possible[:]

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(selected)

    extractors = {
        "Matricule": lambda b: b.employee.matricule,
        "Nom": lambda b: b.employee.name,
        "Departement": lambda b: b.employee.department,
        "TypePrime": lambda b: b.bonus_type.value,
        "DateDebut": lambda b: b.start_date.isoformat(),
        "DateFin": lambda b: b.end_date.isoformat(),
        "Montant": lambda b: str(int(b.total_amount)),
        "Statut": lambda b: b.status.value,
        "DejaRejete": lambda b: "Oui" if b.was_rejected else "Non",
        "CreePar": lambda b: b.created_by.name if b.created_by else '',
        "DateCreation": lambda b: b.created_at.isoformat() if b.created_at else '',
        "Score": lambda b: str(b.performance_score or ''),
        "Quantitatif": lambda b: str(sum(i.get('evaluation', 0) for i in (b.details or {}).get('quantitative', []))),
        "Qualitatif": lambda b: str(sum(i.get('evaluation', 0) for i in (b.details or {}).get('qualitative', []))),
        "NbDisponibilite": lambda b: str(sum(int(d.get('nombre', 0)) for d in (b.details or {}).get('disponibilites', []))),
        "PrimeMaxSemaine": lambda b: str((b.details or {}).get('weekly_max', '')),
        "TauxIntervention": lambda b: str((b.details or {}).get('intervention_rate', '')),
        "TotalDisponibilite": lambda b: str((b.details or {}).get('total_dispo', '')),
        "TotalInterventions": lambda b: str((b.details or {}).get('total_interv', '')),
        "Exceptionnelle": lambda b: str((b.details or {}).get('exceptionnelle', '')),
        "Ponctuelle": lambda b: str((b.details or {}).get('ponctuelle', '')),
        "NbVentes": lambda b: str(sum(int(s.get('nombre', 0)) for s in (b.details or {}).get('sales', []))),
    }

    row = [extractors[col](bonus) for col in selected]
    writer.writerow(row)

    output.seek(0)
    return StreamingResponse(
        iter(['\ufeff' + output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=prime_{bonus.id}_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


# Route GET pour une prime spécifique
@router.get("/bonuses/{bonus_id}", response_model=BonusResponse)
async def get_bonus(bonus_id: int, user: User = Depends(get_current_user)):
    bonus = await Bonus.get_or_none(id=bonus_id).prefetch_related('employee')
    if not bonus: raise HTTPException(404, "Bonus not found")

    # Vérifier le département (sauf DG/DRH)
    if not (user.is_dg or user.is_drh):
        if bonus.employee.dept_str != user.department:
            raise HTTPException(status_code=404, detail="Bonus introuvable")

    # Vérifier que le statut est autorisé pour le rôle
    if user.is_dg:
        allowed = {ValidationStatus.EN_ATTENTE_DG}
    elif user.is_drh:
        allowed = {ValidationStatus.VALIDE}
    elif user.is_directeur:
        allowed = {ValidationStatus.EN_ATTENTE_DIRECTEUR}
    elif user.is_validator_n1:
        allowed = {ValidationStatus.INITIALISE}
    else:
        allowed = set()

    if bonus.status not in allowed:
        raise HTTPException(status_code=404, detail="Bonus introuvable")

    return bonus

# Route GET pour l'historique des validations d'une prime
@router.get("/bonuses/{bonus_id}/validations", response_model=List[ValidationResponse])
async def get_bonus_validations(bonus_id: int):
    bonus = await Bonus.get_or_none(id=bonus_id)
    if not bonus: raise HTTPException(404, "Bonus not found")
    validations = await Validation.filter(bonus_id=bonus_id).prefetch_related('validator')
    result = []
    for v in validations:
        result.append({
            "id": v.id,
            "bonus_id": v.bonus_id,
            "validator_id": v.validator_id,
            "validator_name": v.validator.name if v.validator else None,
            "step": v.step,
            "action": v.action,
            "note": v.note,
            "motif_rejet": v.motif_rejet,
            "validated_at": v.validated_at,
        })
    return result

# Route POST pour valider une prime
@router.post("/bonuses/{bonus_id}/validate")
async def validate_bonus(
    bonus_id: int,
    validation: ValidationCreate,
    step: str,
    user: User = Depends(get_current_user)
):
    bonus = await Bonus.get_or_none(id=bonus_id).prefetch_related('employee')
    if not bonus: raise HTTPException(404, "Bonus not found")
    
    if bonus.status == ValidationStatus.VALIDE:
        raise HTTPException(status_code=400, detail="Bonus déjà validé - aucune action possible")
    
    expected_status = {
        "N1": ValidationStatus.INITIALISE,
        "DIRECTEUR": ValidationStatus.EN_ATTENTE_DIRECTEUR,
        "DG": ValidationStatus.EN_ATTENTE_DG,
    }.get(step)
    
    if not expected_status:
        raise HTTPException(400, "Étape de validation invalide")
    
    if bonus.status != expected_status:
        raise HTTPException(
            400,
            f"Action impossible : la prime est au statut '{bonus.status}', "
            f"attendait '{expected_status}' pour l'étape {step}."
        )
    
    # Vérifier que l'utilisateur a le rôle correspondant à l'étape
    allowed_roles = {
        "N1": user.is_validator_n1,
        "DIRECTEUR": user.is_directeur,
        "DG": user.is_dg,
    }
    if not allowed_roles.get(step, False):
        raise HTTPException(status_code=403, detail=f"Vous n'avez pas le rôle requis pour l'étape '{step}'")
    
    # Vérifier le département (sauf DG)
    if not user.is_dg:
        if bonus.employee.dept_str != user.department:
            raise HTTPException(status_code=403, detail="Vous ne pouvez pas valider une prime d'un autre département")
    
    await Validation.create(
        bonus_id=bonus.id,
        validator_id=user.id,
        step=step,
        action=validation.action,
        note=validation.note,
        motif_rejet=validation.motif_rejet,
    )
    
    # Mise à jour du statut selon l'étape et l'action
    if validation.action == "VALIDER":
        bonus.status = {
            "N1": ValidationStatus.EN_ATTENTE_DIRECTEUR,
            "DIRECTEUR": ValidationStatus.EN_ATTENTE_DG,
            "DG": ValidationStatus.VALIDE
        }[step]
        
        # Clôture automatique si DG valide
        if bonus.status == ValidationStatus.VALIDE:
            await Validation.create(
                bonus_id=bonus.id,
                validator_id=user.id,
                step="CLOSED",
                action="AUTOMATIC",
                note="Prime validée par DG - Clôture automatique"
            )
    elif validation.action == "REJETER":
        bonus.status = ValidationStatus.INITIALISE
        bonus.was_rejected = True
        await AuditLog.create(
            bonus_id=bonus.id,
            user_id=user.id,
            action="REJET",
            description=f"Rejet par {user.name} ({step})" + (f" — Motif : {validation.motif_rejet}" if validation.motif_rejet else ""),
        )

        # Notification au N+1 (manager) que la prime revient à Initialisé
        try:
            employee = await Employee.get(id=bonus.employee_id).prefetch_related('manager')
            if employee.manager:
                motif = f" — Motif : {validation.motif_rejet}" if validation.motif_rejet else ""
                await Notification.create(
                    user=employee.manager,
                    bonus=bonus,
                    sender=user,
                    type="REJET",
                    message=f"{employee.name} — Prime rejetée{motif}",
                )
            else:
                import logging
                logging.warning(f"Pas de manager pour l'employé {employee.name} (id={employee.id})")
        except Exception as e:
            import logging
            logging.error(f"Erreur notification rejet: {e}")

    # Sauvegarde de la prime
    await bonus.save()
    return {"message": "OK", "status": bonus.status}


# Route POST pour marquer des primes comme payées
@router.post("/bonuses/mark-paid")
async def mark_bonuses_paid(req: MarkPaidRequest, user: User = Depends(get_current_user)):
    if not user.is_drh and not user.is_dg:
        raise HTTPException(403, "Seul le DRH peut marquer les primes comme payées")

    query = Bonus.filter(status=ValidationStatus.VALIDE, paid_at__isnull=True)

    if req.bonus_ids:
        query = query.filter(id__in=req.bonus_ids)
    elif req.month and req.year:
        query = query.filter(
            start_date__year=int(req.year),
            start_date__month=int(req.month),
        )
    else:
        raise HTTPException(400, "Fournir bonus_ids ou month+year")

    bonuses = await query
    if not bonuses:
        raise HTTPException(404, "Aucune prime validée trouvée à marquer comme payée")

    now = datetime.utcnow()
    for bonus in bonuses:
        bonus.paid_at = now
        await bonus.save()
        await AuditLog.create(
            bonus_id=bonus.id,
            user_id=user.id,
            action="PAIEMENT",
            description=f"Paiement effectué par {user.name}",
        )

    return {"message": f"{len(bonuses)} prime(s) marquée(s) comme payée(s)", "count": len(bonuses)}

@router.get("/bonuses/{bonus_id}/audit-logs", response_model=List[AuditLogResponse])
async def get_audit_logs(bonus_id: int, user: User = Depends(get_current_user)):
    logs = await AuditLog.filter(bonus_id=bonus_id).prefetch_related('user').order_by('created_at')
    result = []
    for log in logs:
        result.append(AuditLogResponse(
            id=log.id,
            bonus_id=log.bonus_id,
            user_id=log.user_id,
            user_name=log.user.name if log.user else None,
            action=log.action,
            description=log.description,
            changes=log.changes,
            created_at=log.created_at,
        ))
    return result
